import openpyxl
import random
from openpyxl.styles import PatternFill

# Load your existing Excel template
file_path = "Engineering Yield Tracker (1).xlsx"   # <-- change this to your file
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Define columns for yield data (G to P = 10 columns)
data_cols = range(7, 17)  # G=7 → P=16

# Define red fill for cells below target
red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

for row in ws.iter_rows(min_row=4, max_col=6):  # start at row 4 (after headers)
    target_cell = row[3]  # column E = Target Yield
    target_val = target_cell.value

    if target_val in (None, "TBC"):
        continue

    try:
        target_num = float(str(target_val).replace("%", ""))
    except:
        continue

    for col in data_cols:
        # Random variation: ±5% from target
        yield_val = round(random.uniform(target_num - 5, target_num + 3), 2)

        # Ensure values stay within 85–100%
        yield_val = max(85, min(100, yield_val))

        cell = ws.cell(row=row[0].row, column=col)
        cell.value = f"{yield_val}%"

        # Highlight in red if below target
        if yield_val < target_num:
            cell.fill = red_fill

# Save into new file
output_file = "yield_tracking_filled.xlsx"
wb.save(output_file)
print(f"✅ Random yield data written to {output_file}")